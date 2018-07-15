from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import pprint

#load the data file

wb = load_workbook('upazila_indicators.xlsx')


# grab the active worksheet
ws = wb.active

#Declare the required Dictionaries following given attributes
UpazilaData={}


Location={}

PopulationData={}
AllPopulationData={}
PopulationDataBetween06years={}
PopulationDataBetween7to14years={}
PopulationDataBetween15to64years={}
PopulationDataAbove65years={}
PopulationBottomNumber={}


PovertyData={}
ExtremePovertyData={}

Poverty={}

LiteracyData={}
LiteracyAverageData={}
LessThanPrimaryEducation={}
PrimaryEducation={}
SecondaryEducation={}
UniversityEducation={}


AttendanceData={}
SchoolAttendanceof6to18={}
SchoolAttendanceof6to10={}
SchoolAttendanceof11to13={}
SchoolAttendanceof14to15={}
SchoolAttendanceof16to18={}




EmploymentData={}
AgricultureData={}
IndustryData={}
ServiceData={}

HouseholdData={}
ElectricityData={}
ToiletData={}
FlushToiletData={}
NonFlushToiletData={}
WithoutToiletData={}
WaterData={}
TapwaterData={}
TubewellData={}



ChildHealthData={}
UnderWeightChildData={}
SeverelyUnderWeightChildData={}
StuntedChildData={}
SeverelyStuntedChildData={}


#open the file to write the sorted data
print('Writing results...')
resultFile = open('Upaziladata.py', 'w')
resultFile.write('import json\n')
resultFile.write('allData = '+'[')

#loop through each row and write in the file
#I hardcoded the column numbers ; could loop through and collect but the related data aren't always distributed consecutively

for row in range( 2,ws.max_row+1):

    Division  = ws['A' + str(row)].value
    Zila= ws['B' + str(row)].value
    Upazila   = ws['C' + str(row)].value
    
    TotalPopulation =  ws['D' + str(row)].value
    RuralPopulation =  ws['E' + str(row)].value
    WorkingPopulation =  ws['F' + str(row)].value

    PopulationBetween06yearsNumber= ws['DD' + str(row)].value
    PopulationBetween06yearsPercentage= ws['DE' + str(row)].value
    PopulationBetween06yearsNationalAvg=  ws['DF' + str(row)].value
    PopulationBetween06yearsRank= ws['DG' + str(row)].value

    PopulationBetween7to14yearsNumber= ws['DH' + str(row)].value
    PopulationBetween7to14yearsPercentage= ws['DI' + str(row)].value
    PopulationBetween7to14yearsNationalAvg= ws['DJ' + str(row)].value
    PopulationBetween7to14yearsRank= ws['DK' + str(row)].value

    PopulationBetween15to64yearsNumber= ws['DL' + str(row)].value
    PopulationBetween15to64yearsPercentage= ws['DM' + str(row)].value
    PopulationBetween15to64yearsNationalAvg= ws['DN' + str(row)].value
    PopulationBetween15to64yearsRank= ws['DO' + str(row)].value

    PopulationAbove65yearsNumber= ws['DP' + str(row)].value
    PopulationAbove65yearsPercentage= ws['DQ' + str(row)].value
    PopulationAbove65yearsNationalAvg= ws['DR' + str(row)].value
    PopulationAbove65yearsRank= ws['DS' + str(row)].value


    PopulationinBottomNumber= ws['CB' + str(row)].value
    PopulationinBottomPercentage= ws['CC' + str(row)].value
    PopulationinBottomNationalAvg= ws['CD' + str(row)].value
    PopulationinBottomRank= ws['CE' + str(row)].value

    PoorNumber=ws['H' + str(row)].value
    PovertyHeadcountRatio=ws['I' + str(row)].value
    PovertyHeadcountRatioNationalAvg=ws['J' + str(row)].value
    PovertyHeadcountRationRank= ws['K' + str(row)].value


    ExtremePoorNumber=ws['L' + str(row)].value
    ExtremePovertyHeadcountRatio=ws['M' + str(row)].value
    ExtremePovertyHeadcountRatioNationalAvg=ws['N' + str(row)].value
    ExtremePovertyHeadcountRationRank= ws['O' + str(row)].value

    LiteratePopulationNumber= ws['AN' + str(row)].value
    LiteratePopulationPercentage= ws['AO' + str(row)].value
    LiteratePopulationNationalAvg= ws['AP' + str(row)].value
    LiteratePopulationRank= ws['AQ' + str(row)].value

    LessThanPrimaryEducationNumber=  ws['AN' + str(row)].value
    LessThanPrimaryEducationPercentage= ws['AN' + str(row)].value
    LessThanPrimaryEducationNationalAvg= ws['AN' + str(row)].value
    LessThanPrimaryEducationRank= ws['AN' + str(row)].value

    PrimaryEducationNumber= ws['AN' + str(row)].value
    PrimaryEducationPercentage= ws['AN' + str(row)].value
    PrimaryEducationNationalAvg= ws['AN' + str(row)].value
    PrimaryEducationRank= ws['AN' + str(row)].value

    

    SecondaryEducationNumber= ws['AN' + str(row)].value
    SecondaryEducationPercentage= ws['AN' + str(row)].value
    SecondaryEducationNationalAvg= ws['AN' + str(row)].value
    SecondaryEducationRank= ws['AN' + str(row)].value

 
    UniversityEducationNumber= ws['AN' + str(row)].value
    UniversityEducationPercentage= ws['AN' + str(row)].value
    UniversityEducationNationalAvg= ws['AN' + str(row)].value
    UniversityEducationRank= ws['AN' + str(row)].value

    SchoolAttendanceAmong6to18YearsNumber= ws['BH' + str(row)].value
    SchoolAttendanceAmong6to18YearsPercentage=ws['BI' + str(row)].value
    SchoolAttendanceAmong6to18YearsNationalAvg=ws['BJ' + str(row)].value
    SchoolAttendanceAmong6to18YearsRank=ws['BK' + str(row)].value

    SchoolAttendanceAmong6to10YearsNumber=ws['BL' + str(row)].value
    SchoolAttendanceAmong6to10YearsPercentage=ws['BM' + str(row)].value
    SchoolAttendanceAmong6to10YearsNationalAvg=ws['BN' + str(row)].value
    SchoolAttendanceAmong6to10YearsRank=ws['BO' + str(row)].value
    
    SchoolAttendanceAmong11to13YearsNumber=ws['BP' + str(row)].value
    SchoolAttendanceAmong11to13YearsPercentage=ws['BQ' + str(row)].value
    SchoolAttendanceAmong11to13YearsNationalAvg=ws['BR' + str(row)].value
    SchoolAttendanceAmong11to13YearsRank=ws['BS' + str(row)].value

    SchoolAttendanceAmong14to15YearsNumber=ws['BT' + str(row)].value
    SchoolAttendanceAmong14to15YearsPercentage=ws['BU' + str(row)].value
    SchoolAttendanceAmong14to15YearsNationalAvg=ws['BV' + str(row)].value
    SchoolAttendanceAmong14to15YearsRank=ws['BX' + str(row)].value

    SchoolAttendanceof16to18YearsNumber=ws['BW' + str(row)].value
    SchoolAttendanceof16to18YearsPercentage=ws['BX' + str(row)].value
    SchoolAttendanceof16to18YearsNationalAvg=ws['BY' + str(row)].value
    SchoolAttendanceof16to18YearsRank=ws['BZ' + str(row)].value

    
    AgricultureNumber=ws['P' + str(row)].value
    AgriculturePercent=ws['Q' + str(row)].value
    AgricultureNationalAvg=ws['R' + str(row)].value
    AgricultureRank=ws['S' + str(row)].value

    IndustryNumber=ws['T' + str(row)].value
    IndustryPercent=ws['U' + str(row)].value
    IndustryNationalAvg=ws['V' + str(row)].value
    IndustryRank=ws['W' + str(row)].value

    ServicesNumber=ws['X' + str(row)].value
    ServicesPercentage=ws['Y' + str(row)].value
    ServicesNationalAvg=ws['Z' + str(row)].value
    ServicesRank=ws['AA' + str(row)].value

    NumberOfHousehold= ws['G' + str(row)].value

    ElectricityNumber=ws['AB' + str(row)].value
    ElectricityPercentage=ws['AC' + str(row)].value
    ElectricityNationalAvg=ws['AD' + str(row)].value
    ElectricityRank=ws['AE' + str(row)].value

    FlushToiletNumber=ws['AF' + str(row)].value
    FlushToiletPercentage=ws['AG' + str(row)].value
    FlushToiletNationalAvg=ws['AH' + str(row)].value
    FlushToiletRank=ws['AI' + str(row)].value

    NonFlushToiletNumber=ws['AJ' + str(row)].value
    NonFlushToiletPercentage=ws['AK' + str(row)].value
    NonFlushToiletNationalAvg=ws['AL' + str(row)].value
    NonFlushToiletRank=ws['AM' + str(row)].value

    HouseholdswithoutToiletNumber=ws['DT' + str(row)].value
    HouseholdswithoutToiletPercentage=ws['DU' + str(row)].value
    HouseholdswithoutToiletNationalAvg=ws['DV' + str(row)].value
    HouseholdswithoutToiletRank=ws['DW' + str(row)].value
    
    HouseholdsWithTapwaterNumber=ws['CV' + str(row)].value
    HouseholdsWithTapwaterPercentage=ws['CW' + str(row)].value
    HouseholdsWithTapwaterNationalAvg=ws['CX' + str(row)].value
    HouseholdsWithTapwaterRank=ws['CY' + str(row)].value
    
    
    HouseholdsWithtubewellNumber=ws['CZ' + str(row)].value
    HouseholdsWithtubewellPercentage=ws['DA' + str(row)].value
    HouseholdsWithtubewellNationalAvg=ws['DB' + str(row)].value
    HouseholdsWithtubewellRank=ws['DC' + str(row)].value

       

    underweightChildrenNumber=ws['CF' + str(row)].value
    underweightChildrenPercentage=ws['CG' + str(row)].value
    underweightChildrenNationalAvg=ws['CH' + str(row)].value
    underweightChildrenRank=ws['CI' + str(row)].value

    
    severelyunderweightChildrenNumber=ws['CJ' + str(row)].value
    severelyunderweightChildrenPercentage=ws['CK' + str(row)].value
    severelyunderweightChildrenNationalAvg=ws['CL' + str(row)].value
    severelyunderweightChildrenRank=ws['CM' + str(row)].value

    stuntedChildrenNumber=ws['CN' + str(row)].value
    stuntedChildrenPercentage=ws['CO' + str(row)].value
    stuntedChildrenNationalAvg=ws['CP' + str(row)].value
    stuntedChildrenRank=ws['CQ' + str(row)].value
    
    severelystuntedChildrenNumber=ws['CR' + str(row)].value
    severelystuntedChildrenPercentage=ws['CS' + str(row)].value
    severelystuntedChildrenNationalAvg=ws['CT' + str(row)].value
    severelystuntedChildrenRank=ws['CU' + str(row)].value


#collect into separate dictionary for each attribute
#location info
    LocationData={
        'division': Division,
         'zila':  Zila,
        'upazila': Upazila
        }
   

#population related data
    AllPopulationData={
        'total':TotalPopulation,
        'rural':RuralPopulation,
        'working':WorkingPopulation
        }
   
    PopulationDataBetween06years={
          'number': PopulationBetween06yearsNumber,
          'percentage': PopulationBetween06yearsPercentage,
          'nationalAverage' :PopulationBetween06yearsNationalAvg,
          'rank': PopulationBetween06yearsRank
        }
    PopulationDataBetween7to14years={
          'number': PopulationBetween7to14yearsNumber,
          'percentage': PopulationBetween7to14yearsPercentage,
          'nationalAverage' :PopulationBetween7to14yearsNationalAvg,
          'rank': PopulationBetween7to14yearsRank
        }
    
    PopulationDataBetween15to64years={
          'number':  PopulationBetween15to64yearsNumber,
          'percentage':  PopulationBetween15to64yearsPercentage,
          'nationalAverage' : PopulationBetween15to64yearsNationalAvg,
          'rank':  PopulationBetween15to64yearsRank
        }
    PopulationDataAbove65years={
          'number':  PopulationAbove65yearsNumber,
          'percentage':  PopulationAbove65yearsPercentage,
          'nationalAverage' : PopulationAbove65yearsNationalAvg,
          'rank':  PopulationAbove65yearsRank

        }
    PopulationBottomNumber={
         'number':  PopulationinBottomNumber,
          'percentage':  PopulationinBottomPercentage,
          'nationalAverage' : PopulationinBottomNationalAvg,
          'rank':  PopulationinBottomRank

        }
    
    PopulationData={
        'summary': AllPopulationData,
        'between_0_6 _years': PopulationDataBetween06years,
        'between_7_14_years': PopulationDataBetween7to14years,
        'between_15_64_years':PopulationDataBetween15to64years,
        'above_65_years':PopulationDataAbove65years,
        'bottom_population': PopulationBottomNumber,
         }
   
#poverty hit data
    ExtremePovertyData={
         'number':  ExtremePoorNumber,
          'percentage':  ExtremePovertyHeadcountRatio,
          'nationalAverage' : ExtremePovertyHeadcountRatioNationalAvg,
          'rank':  ExtremePovertyHeadcountRationRank
        }
   

    PovertyData={
         'number':  PoorNumber,
          'percentage':  PovertyHeadcountRatio,
          'nationalAverage' : PovertyHeadcountRatioNationalAvg,
          'rank':  PovertyHeadcountRationRank
        }
    TotalPovertyData={
        'poverty':PovertyData,
        'extreme_poverty_data': ExtremePovertyData
        }
   

#Education level    
    LessThanPrimaryEducation={
        'number':  LessThanPrimaryEducationNumber,
          'percentage':  LessThanPrimaryEducationPercentage,
          'nationalAverage' : LessThanPrimaryEducationNationalAvg,
          'rank':  LessThanPrimaryEducationRank


        }
    PrimaryEducation={
          'number':  PrimaryEducationNumber,
          'percentage':  PrimaryEducationPercentage,
          'nationalAverage' : PrimaryEducationNationalAvg,
          'rank':  PrimaryEducationRank
          }
    SecondaryEducation={
          'mumber':  SecondaryEducationNumber,
          'percentage':  SecondaryEducationPercentage,
          'nationalAverage' : SecondaryEducationNationalAvg,
          'rank':  SecondaryEducationRank
          }
    UniversityEducation={
          'number':  UniversityEducationNumber,
          'percentage':  UniversityEducationPercentage,
          'nationalAverage' : UniversityEducationNationalAvg,
          'rank':  UniversityEducationRank
          }
    
    LiteracyAverageData={
        'number':  LiteratePopulationNumber,
          'percentage':  LiteratePopulationPercentage,
          'nationalAverage' : LiteratePopulationNationalAvg,
          'rank':  LiteratePopulationRank
        }

    LiteracyData={
          'less_than_primary_education':LessThanPrimaryEducation,
          'primary_education': PrimaryEducation,
          'secondary_ education':SecondaryEducation,
          'university_education':UniversityEducation,
          'average_data': LiteracyAverageData

        }
    
    
    TotalLiteracyData={
        'literacy': LiteracyData
        }
    
        
#school attendance data    
    SchoolAttendanceof6to18={
        'number':SchoolAttendanceAmong6to18YearsNumber,
        'percentage':SchoolAttendanceAmong6to18YearsPercentage,
        'national Avg':SchoolAttendanceAmong6to18YearsNationalAvg,
        'rank': SchoolAttendanceAmong6to18YearsRank
        }
    SchoolAttendanceof6to10={
        'number':SchoolAttendanceAmong6to10YearsNumber,
        'percentage':SchoolAttendanceAmong6to10YearsPercentage,
        'national Avg':SchoolAttendanceAmong6to10YearsNationalAvg,
        'rank':SchoolAttendanceAmong6to10YearsRank
        }
    SchoolAttendanceof11to13={
        'number': SchoolAttendanceAmong11to13YearsNumber,
        'percentage': SchoolAttendanceAmong11to13YearsPercentage,
        'national Avg': SchoolAttendanceAmong11to13YearsNationalAvg,
        'rank':  SchoolAttendanceAmong11to13YearsRank}
    SchoolAttendanceof14to15={
       'number': SchoolAttendanceAmong14to15YearsNumber,
       'percentage': SchoolAttendanceAmong14to15YearsPercentage,
         'national Avg':SchoolAttendanceAmong14to15YearsNationalAvg,
        'rank':SchoolAttendanceAmong14to15YearsRank
        }
    SchoolAttendanceof16to18={
        'number':SchoolAttendanceof16to18YearsNumber,
        'percentage':SchoolAttendanceof16to18YearsPercentage,
         'national Avg':SchoolAttendanceof16to18YearsNationalAvg,
        'rank':SchoolAttendanceof16to18YearsRank
        }
    AttendanceData={
        'school_attendance_of_6_to_18_years': SchoolAttendanceof6to18,
        'school_attendance_of_6_to_10_years': SchoolAttendanceof6to10,
        'school_attendance_of_11_to_13_years': SchoolAttendanceof11to13,
        'school-attendance_of_14_to_15_years': SchoolAttendanceof14to15,
        'school_attendance_of_16_to_18_years': SchoolAttendanceof16to18
        }

    
#employment data
    
    AgricultureData={
        'number':AgricultureNumber,
        'percentage':AgriculturePercent,
        'national Avg':AgricultureNationalAvg,
        'rank':AgricultureRank
        }
    IndustryData={
        'number':IndustryNumber,
        'percentage':IndustryPercent,
        'national Avg':IndustryNationalAvg,
        'rank':IndustryRank
        }
    ServiceData={
        'number':ServicesNumber,
        'percentage':ServicesPercentage,
        'national_avg':ServicesNationalAvg,
        'rank':ServicesRank

        }
    EmploymentData={
        'agriculture_employment_info': AgricultureData,
        'industrial_employment_info': IndustryData,
        'service_holders': ServiceData
        }

  #household data  
    ElectricityData={
        'number':ElectricityNumber,
        'percentage':ElectricityPercentage,
        'national_avg':ElectricityNationalAvg,
        'rank':ElectricityRank
        }
    
   
    
    FlushToiletData={
        'number':FlushToiletNumber,
        'percentage':FlushToiletPercentage,
        'national_avg':FlushToiletNationalAvg,
        'rank':FlushToiletRank
        }
    NonFlushToiletData={
         'number':NonFlushToiletNumber,
        'percentage':NonFlushToiletPercentage,
        'national_avg':NonFlushToiletNationalAvg,
        'rank':NonFlushToiletRank
         }
    WithoutToiletData={
         'number':HouseholdswithoutToiletNumber,
        'percentage':HouseholdswithoutToiletPercentage,
        'national-Avg':HouseholdswithoutToiletNationalAvg,
        'rank':HouseholdswithoutToiletRank
    }
    
    ToiletData={
        'flush_toilet':FlushToiletData,
        'without_flush_toilet':NonFlushToiletData,
        'without_toilet':WithoutToiletData
        }

   
    TapwaterData={
        'number':HouseholdsWithTapwaterNumber,
        'percentage':HouseholdsWithTapwaterPercentage,
        'national Avg':HouseholdsWithTapwaterNationalAvg,
        'rank':HouseholdsWithTapwaterRank
            
        }
    TubewellData={
        'number':HouseholdsWithtubewellNumber,
        'percentage':HouseholdsWithtubewellPercentage,
        'national_Avg':HouseholdsWithtubewellNationalAvg,
        'rank':HouseholdsWithtubewellRank
        }

    WaterData={
        'tapwater': TapwaterData,
        'tubewell_water': TubewellData
        }

    
    HouseholdData={
         'number_of_household':NumberOfHousehold,
         'electricity':ElectricityData,
         'toilet':ToiletData,
         'water': WaterData
         
        }
  

  

    #child health data
    UnderWeightChildData={
            'number':  underweightChildrenNumber,
          'percentage': underweightChildrenPercentage,
          'nationalAverage' : underweightChildrenNationalAvg,
          'rank': underweightChildrenRank
        }
    SeverelyUnderWeightChildData={
            'number': severelyunderweightChildrenNumber,
          'percentage': severelyunderweightChildrenPercentage,
          'nationalAverage' : severelyunderweightChildrenNationalAvg,
          'rank':  severelyunderweightChildrenRank
        }
    StuntedChildData={
            'number':  stuntedChildrenNumber,
           'percentage': stuntedChildrenPercentage,
          'nationalAverage' : stuntedChildrenNationalAvg,
          'rank':  stuntedChildrenRank
        }
    SeverelyStuntedChildData={
            'number':  severelystuntedChildrenNumber,
           'percentage': severelystuntedChildrenPercentage,
          'nationalAverage' : severelystuntedChildrenNationalAvg,
          'rank':  severelystuntedChildrenRank
        }
    ChildHealthData={
        'underweight_children':UnderWeightChildData,
        'severely_underweight_children':SeverelyUnderWeightChildData,
        'stunted_children':StuntedChildData,
        'severely_stunted_children':SeverelyStuntedChildData
        }
    
    #finally, all the attributes under the hood of 8 main attributes
    UpazilaData= {
        '_id:' : str(row-1),
       'location_data': LocationData,
       'population_data': PopulationData,
       'poverty_data': TotalPovertyData,
       'literate_data': LiteracyData,
       'attendance_data': AttendanceData,
       'employment_data': EmploymentData,
       'household_info': HouseholdData,
       'child_health_data': ChildHealthData
       
    }

    resultFile.write(pprint.pformat(UpazilaData))
    resultFile.write(',\n')

resultFile.write(']')   
resultFile.close()
print('Done.')

wb.template = False
wb.save('final.xlsx')

